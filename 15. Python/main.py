import openpyxl

# Task:
#     - List each company with respective product count
#     - List product with inventory less than 10
#     - List each company with respective total inventory value
#     - Write inventory value for each product into spreadsheet

inv_file = openpyxl.load_workbook("inventory.xlsx")
sheet1 = inv_file["Sheet1"]

products_per_supplier = {}
products_less_than_10 = []
total_value_per_supplier = {}

for row in range(2,sheet1.max_row + 1):
    supplier_name = sheet1.cell(row,4).value

    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] += 1
    else:
        products_per_supplier[supplier_name] = 1

    product_id = sheet1.cell(row, 1).value
    inventory = int(sheet1.cell(row, 2).value)

    if (inventory < 10):
        products_less_than_10.append(product_id)

    price = sheet1.cell(row, 3).value

    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] = total_value_per_supplier[supplier_name] + (inventory * price)
    else:
        total_value_per_supplier[supplier_name] = (inventory * price)

    # Write inventory value of each product into sheet
    total_value = inventory * price
    cell_object = sheet1.cell(row,5)
    cell_object.value = total_value

print(products_per_supplier)
print(products_less_than_10)
print(total_value_per_supplier)
inv_file.save("inventory_with_total_value.xlsx") 